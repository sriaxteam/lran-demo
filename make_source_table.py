# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "수집 소스 목록"

# ── 색상 정의
COLOR = {
    "header_bg":   "1B3A5C",
    "header_font": "FFFFFF",
    "cat1_bg":     "1B3A5C",   # 언론
    "cat2_bg":     "2E4E7E",   # 싱크탱크
    "cat3_bg":     "1E6B8A",   # 국제기구
    "cat4_bg":     "1A6B4A",   # 국내 연구기관
    "cat5_bg":     "5A3E7A",   # 정책 벤치마킹
    "cat_font":    "FFFFFF",
    "row_odd":     "F8FAFC",
    "row_even":    "FFFFFF",
    "border":      "D1D9E0",
}

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def border():
    s = Side(style="thin", color=COLOR["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def hfont(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

# ── 데이터
SOURCES = [
    # (구분, 소스명, 공신력, 웹사이트, RSS/수집방식, 선별 이유)
    ("언론", "Reuters",        "9.6", "https://www.reuters.com/world",                          "https://feeds.reuters.com/reuters/worldNews",                                                    "세계 최고 공신력, 에너지·군사 속보 최강"),
    ("언론", "AP News",        "9.5", "https://apnews.com/world-news",                          "https://feeds.apnews.com/rss/apf-worldnews",                                                    "Reuters 백업, 팩트체크 기준"),
    ("언론", "BBC 중동",       "9.0", "https://www.bbc.com/news/world/middle_east",              "https://feeds.bbci.co.uk/news/world/middle_east/rss.xml",                                       "중동 전담 데스크, 현장 밀착 보도"),
    ("언론", "Bloomberg",      "9.0", "https://www.bloomberg.com/markets",                      "https://feeds.bloomberg.com/markets/news.rss",                                                  "유가·LNG·에너지 시장 실시간 수치"),
    ("언론", "연합뉴스",       "8.5", "https://www.yna.co.kr/international",                    "https://www.yna.co.kr/rss/international.xml",                                                   "한국어 속보, 한국 정부 반응 가장 빠름"),
    ("언론", "FT World",       "9.1", "https://www.ft.com/world",                               "https://www.ft.com/world?format=rss",                                                           "에너지·경제 분석 깊이 최상급"),
    ("언론", "Al Jazeera",     "8.0", "https://www.aljazeera.com",                              "https://www.aljazeera.com/xml/rss/all.xml",                                                     "아랍·이란 시각, 서방 언론과 균형"),
    ("언론", "Guardian 이란",  "8.7", "https://www.theguardian.com/world/iran",                 "https://www.theguardian.com/world/iran/rss",                                                    "이란 전담 섹션, 인도주의·외교 커버리지"),
    ("언론", "NYT World",      "8.8", "https://www.nytimes.com/section/world",                  "https://rss.nytimes.com/services/xml/rss/nyt/World.xml",                                        "미국 행정부·의회 반응 파악 필수"),
    ("언론", "SCMP",           "8.0", "https://www.scmp.com/news/world",                        "https://www.scmp.com/rss/91/feed",                                                              "중국 시각, 이란 원유 수입·대미 관계 분석용"),

    ("싱크탱크·정책연구", "Foreign Affairs", "9.0", "https://www.foreignaffairs.com",                        "https://www.foreignaffairs.com/rss.xml",                                                        "가장 권위 있는 국제관계 분석지"),
    ("싱크탱크·정책연구", "CSIS",            "8.8", "https://www.csis.org/programs/latest-analysis-war-iran","https://www.csis.org/rss.xml",                                                                 "미국 對이란 정책 분석 핵심 싱크탱크"),
    ("싱크탱크·정책연구", "CFR 이란",        "8.6", "https://www.cfr.org/region/iran",                       "https://www.cfr.org/rss.xml",                                                                  "이란 전용 분석 페이지, 배경·맥락 최강"),
    ("싱크탱크·정책연구", "Al-Monitor",      "8.6", "https://www.al-monitor.com",                            "https://www.al-monitor.com/rss",                                                               "이란·중동 현지 전문 미디어, 타 언론 미커버 내용"),
    ("싱크탱크·정책연구", "War on the Rocks","8.5", "https://warontherocks.com",                             "https://warontherocks.com/feed/",                                                              "군사전략 심층분석, 호르무즈 봉쇄 시나리오"),

    ("국제기구", "IEA",    "9.8", "https://www.iea.org/topics/oil-markets",                "Google News RSS: news.google.com/rss/search?q=site:iea.org+energy",          "국제 에너지 공식 수치, 보고서 인용 최다"),
    ("국제기구", "IMF",    "9.8", "https://www.imf.org/en/Publications/WEO",               "Google News RSS: news.google.com/rss/search?q=site:imf.org",                 "한국 성장률 하향·물가 전망 공식 근거"),
    ("국제기구", "OPEC",   "9.5", "https://www.opec.org/opec_web/en/publications/338.htm", "Google News RSS: news.google.com/rss/search?q=site:opec.org+oil",           "감산·증산 결정, 유가 방향 직결"),
    ("국제기구", "UNCTAD", "9.5", "https://unctad.org/topic/trade-analysis",               "Google News RSS: news.google.com/rss/search?q=site:unctad.org+trade",       "호르무즈·수에즈 물동량·해운 공식 통계"),

    ("국내 연구기관", "KEEI (한국에너지경제연구원)", "9.0", "https://www.keei.re.kr", "https://www.keei.re.kr/keei/rss/newsRss.rss",    "한국 에너지비 분석 유일한 국내 공식 기관"),
    ("국내 연구기관", "KIEP (대외경제정책연구원)",   "9.0", "https://www.kiep.go.kr", "https://www.kiep.go.kr/rss/rssFeed.do",          "이란전쟁→한국 수출·공급망 파급 효과 논거"),
    ("국내 연구기관", "KDI (한국개발연구원)",        "9.0", "https://www.kdi.re.kr",  "https://www.kdi.re.kr/common/rss.jsp",           "물가·고용·소상공인 국내 거시경제 기준"),

    ("정책 벤치마킹", "EU 에너지 정책", "9.0", "https://energy.ec.europa.eu/news_en",  "https://ec.europa.eu/rss/energy_en.xml",   "에너지 바우처·가격 상한제 직접 모델"),
    ("정책 벤치마킹", "일본 METI",      "9.0", "https://www.meti.go.jp/english",        "RSS 없음 — 직접 스크래핑",                 "LNG 비축·에너지 보조금 직접 지급 선례"),
    ("정책 벤치마킹", "DefenseOne",     "8.0", "https://www.defenseone.com",            "https://www.defenseone.com/rss/all/",      "호르무즈 군사 작전 현황, 봉쇄 시나리오 판단용"),
]

CAT_COLORS = {
    "언론":             "1B3A5C",
    "싱크탱크·정책연구":"2E4E7E",
    "국제기구":         "1E6B8A",
    "국내 연구기관":    "1A6B4A",
    "정책 벤치마킹":    "5A3E7A",
}

# ── 제목
ws.merge_cells("A1:F1")
ws["A1"] = "이란전쟁 민생 에이전트 — 분석 소스 목록 (25개 선별)"
ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws["A1"].fill = fill("0D1B2A")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:F2")
ws["A2"] = "수원시정연구원 | 이란-미 전쟁 민생 이슈 발굴 에이전트 v2.0 | 선별 기준: 속보성 / 중동 전문성 / 에너지·경제 커버리지 / 한국 연결성 / 중복 제거"
ws["A2"].font = Font(name="Arial", size=9, color="94A3B8")
ws["A2"].fill = fill("0D1B2A")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# ── 헤더
headers = ["구분", "소스명", "공신력", "웹사이트 (직접 방문)", "RSS / 수집 방식", "선별 이유"]
ws.append(headers)
header_row = ws.max_row
for col, _ in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=col)
    cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell.fill = fill("1B3A5C")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border()
ws.row_dimensions[header_row].height = 22

# ── 데이터 행
prev_cat = None
row_in_cat = 0

for i, (cat, name, cred, web, rss, reason) in enumerate(SOURCES):
    if cat != prev_cat:
        row_in_cat = 0
        prev_cat = cat
    row_in_cat += 1

    ws.append([cat, name, cred, web, rss, reason])
    r = ws.max_row
    ws.row_dimensions[r].height = 20

    cat_color = CAT_COLORS.get(cat, "374151")
    bg = "F8FAFC" if row_in_cat % 2 == 1 else "FFFFFF"

    for col in range(1, 7):
        cell = ws.cell(row=r, column=col)
        cell.border = border()
        cell.alignment = Alignment(vertical="center", wrap_text=True,
                                   horizontal="center" if col in [1, 3] else "left")
        if col == 1:
            cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            cell.fill = fill(cat_color)
        elif col == 2:
            cell.font = Font(name="Arial", bold=True, size=9, color="0D1B2A")
            cell.fill = fill(bg)
        elif col == 3:
            cell.font = Font(name="Arial", bold=True, size=9, color="1A6B4A")
            cell.fill = fill(bg)
        elif col in [4, 5]:
            cell.font = Font(name="Arial", size=8, color="1D4ED8")
            cell.fill = fill(bg)
        else:
            cell.font = Font(name="Arial", size=9, color="374151")
            cell.fill = fill(bg)

# ── 열 너비
col_widths = [16, 26, 8, 48, 60, 36]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── 주석 행
ws.append([])
ws.append(["※ 비고", "", "", "", "", ""])
note_row = ws.max_row
ws.cell(note_row, 1).font = Font(name="Arial", bold=True, size=9, color="C0392B")

notes = [
    "① 국제기구(IEA·IMF·OPEC·UNCTAD)는 직접 RSS를 차단하여 Google News RSS 우회 수집",
    "② 일본 METI는 RSS 미제공 — 웹사이트 직접 스크래핑으로 수집",
    "③ 공신력 점수: 10.0 만점 기준 (Reuters 9.6 최고 / AsiaTimes 7.5 최저)",
]
for note in notes:
    ws.append(["", note, "", "", "", ""])
    r = ws.max_row
    ws.cell(r, 2).font = Font(name="Arial", size=9, color="64748B", italic=True)
    ws.row_dimensions[r].height = 16

# ── 저장
out = r"C:\Users\user\Desktop\iran_news_agent_final\iran_final\이란전쟁_에이전트_소스목록.xlsx"
wb.save(out)
print(f"저장 완료: {out}")
