# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = load_workbook(r"C:\Users\user\Desktop\iran_news_agent_final\iran_final\이란전쟁_에이전트_소스목록.xlsx")

# ── 키워드 데이터
KEYWORDS = {
    "영문 키워드 — 이란 직접": {
        "color": "C0392B",
        "desc": "이란 전쟁·군사 행동 직접 관련 필터",
        "items": [
            ("iran war",         "이란 전쟁"),
            ("iran us war",      "이란-미국 전쟁"),
            ("hormuz",           "호르무즈 해협"),
            ("hormuz blockade",  "호르무즈 봉쇄"),
            ("tehran",           "테헤란"),
            ("irgc",             "이란 혁명수비대"),
            ("iranian",          "이란의"),
            ("iran ceasefire",   "이란 휴전"),
            ("iran nuclear",     "이란 핵"),
            ("iran missile",     "이란 미사일"),
            ("hezbollah iran",   "헤즈볼라-이란"),
            ("strait of hormuz", "호르무즈 해협 (전체명)"),
            ("iran oil",         "이란 원유"),
            ("iran blockade",    "이란 봉쇄"),
            ("iran sanction",    "이란 제재"),
            ("iran attack",      "이란 공격"),
            ("iran deal",        "이란 핵합의"),
        ]
    },
    "영문 키워드 — 중동 지역 정세": {
        "color": "E67E22",
        "desc": "중동 분쟁·지역 갈등 간접 모니터링",
        "items": [
            ("middle east war",      "중동 전쟁"),
            ("middle east conflict", "중동 분쟁"),
            ("persian gulf",         "페르시아만"),
            ("gulf crisis",          "걸프 위기"),
            ("red sea",              "홍해"),
            ("houthi",               "후티 반군"),
            ("hezbollah",            "헤즈볼라"),
            ("israel iran",          "이스라엘-이란"),
            ("israel strike",        "이스라엘 공습"),
            ("israel war",           "이스라엘 전쟁"),
            ("gaza war",             "가자 전쟁"),
            ("west bank",            "웨스트뱅크"),
            ("hamas",                "하마스"),
        ]
    },
    "영문 키워드 — 에너지·유가": {
        "color": "D97706",
        "desc": "에너지 가격·공급 영향 — 수원시 도시가스·전기요금 연동",
        "items": [
            ("oil price",        "유가"),
            ("crude oil",        "원유"),
            ("oil market",       "원유 시장"),
            ("opec cut",         "OPEC 감산"),
            ("energy crisis",    "에너지 위기"),
            ("energy security",  "에너지 안보"),
            ("oil supply",       "원유 공급"),
            ("lng price",        "LNG 가격"),
            ("gas price spike",  "가스가격 급등"),
            ("fuel price",       "연료비"),
            ("oil sanction",     "원유 제재"),
            ("energy sanction",  "에너지 제재"),
        ]
    },
    "영문 키워드 — 경제·공급망": {
        "color": "1A6B4A",
        "desc": "공급망·물류 충격 — 수원시 제조업·소상공인 영향",
        "items": [
            ("shipping disruption", "해운 차질"),
            ("suez canal",          "수에즈 운하"),
            ("supply chain",        "공급망"),
            ("tanker attack",       "유조선 공격"),
            ("naval blockade",      "해상 봉쇄"),
            ("oil tanker",          "유조선"),
            ("global inflation",    "글로벌 인플레이션"),
            ("war economy",         "전시 경제"),
        ]
    },
    "한국어 키워드 — 이란 직접": {
        "color": "C0392B",
        "desc": "한국어 뉴스·연구기관 이란 관련 필터",
        "items": [
            ("이란 전쟁", ""),
            ("이란 미국", ""),
            ("호르무즈",  ""),
            ("이란 핵",   ""),
            ("이란 봉쇄", ""),
            ("테헤란",    ""),
            ("이란 휴전", ""),
            ("이란전",    ""),
            ("이란 공격", ""),
        ]
    },
    "한국어 키워드 — 중동 지역": {
        "color": "E67E22",
        "desc": "한국어 뉴스 중동 분쟁 필터",
        "items": [
            ("중동전쟁",    ""),
            ("중동 분쟁",   ""),
            ("페르시아만",  ""),
            ("홍해",        ""),
            ("후티",        ""),
            ("헤즈볼라",    ""),
            ("이스라엘 이란",""),
            ("가자",        ""),
            ("하마스",      ""),
        ]
    },
    "한국어 키워드 — 에너지·유가": {
        "color": "D97706",
        "desc": "한국 도시가스·전기요금 관련 직접 필터",
        "items": [
            ("유가",      ""),
            ("원유",      ""),
            ("에너지 위기",""),
            ("에너지 안보",""),
            ("유류비",    ""),
            ("난방비",    ""),
            ("도시가스",  ""),
            ("LNG 가격",  ""),
        ]
    },
    "한국어 키워드 — 경제·물가": {
        "color": "1A6B4A",
        "desc": "수원시 민생 물가·공급망 직접 연계 필터",
        "items": [
            ("물가",     ""),
            ("인플레",   ""),
            ("공급망",   ""),
            ("해운 운임",""),
        ]
    },
    "민생 특화 키워드 (수원시 직결)": {
        "color": "1B3A5C",
        "desc": "수원시 민생 분석 전용 — 정책 보고서·지역 언론 필터",
        "items": [
            ("유가",     ""),
            ("물가",     ""),
            ("에너지",   ""),
            ("난방비",   ""),
            ("전기료",   ""),
            ("도시가스", ""),
            ("소상공인", ""),
            ("민생",     ""),
            ("장바구니", ""),
            ("생활비",   ""),
            ("인플레",   ""),
            ("수원시",   ""),
            ("경기도",   ""),
            ("취약계층", ""),
        ]
    },
    "패러다임 변화 감지 키워드": {
        "color": "5A3E7A",
        "desc": "에너지·무역 구조 재편 신호 감지 — 장기 시나리오 분석용",
        "items": [
            ("energy security",       "에너지 안보"),
            ("energy transition",     "에너지 전환"),
            ("supply diversification","공급 다변화"),
            ("strategic reserve",     "전략 비축"),
            ("energy partnership",    "에너지 파트너십"),
            ("reliability",           "공급 신뢰성"),
            ("recession",             "경기침체"),
            ("growth forecast",       "성장 전망"),
            ("inflation outlook",     "인플레이션 전망"),
            ("trade disruption",      "무역 차질"),
            ("supply chain",          "공급망"),
            ("shipping route",        "해운 항로"),
            ("price cap",             "가격 상한제"),
            ("export controls",       "수출 통제"),
            ("hoarding",              "사재기"),
            ("emergency",             "비상사태"),
            ("paradigm",              "패러다임"),
            ("fragmentation",         "공급망 분절화"),
            ("geopolitical",          "지정학적"),
            ("decoupling",            "디커플링"),
            ("에너지안보",            ""),
            ("패러다임",              ""),
            ("공급망재편",            ""),
            ("무역질서",              ""),
        ]
    },
}

# ── 시트 생성
ws = wb.create_sheet("키워드 목록")

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def border_thin():
    s = Side(style="thin", color="D1D9E0")
    return Border(left=s, right=s, top=s, bottom=s)

# 제목
ws.merge_cells("A1:E1")
ws["A1"] = "이란전쟁 민생 에이전트 — 수집 키워드 목록"
ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws["A1"].fill = fill("0D1B2A")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:E2")
ws["A2"] = "총 5개 그룹 · 영문 50개 + 한국어 30개 + 민생특화 14개 + 패러다임 24개 = 약 118개 키워드"
ws["A2"].font = Font(name="Arial", size=9, color="94A3B8")
ws["A2"].fill = fill("0D1B2A")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 16

# 헤더
headers = ["그룹", "키워드", "한국어 의미", "용도 설명", "언어"]
ws.append(headers)
hr = ws.max_row
for col in range(1, 6):
    c = ws.cell(hr, col)
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = fill("1B3A5C")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin()
ws.row_dimensions[hr].height = 22

# 데이터
for group_name, gdata in KEYWORDS.items():
    color = gdata["color"]
    desc  = gdata["desc"]
    lang  = "EN" if "영문" in group_name else ("KO" if "한국어" in group_name or "민생" in group_name else "EN/KO")
    items = gdata["items"]

    for idx, (kw, meaning) in enumerate(items):
        ws.append([group_name if idx == 0 else "", kw, meaning, desc if idx == 0 else "", lang])
        r = ws.max_row
        ws.row_dimensions[r].height = 18
        bg = "F8FAFC" if idx % 2 == 0 else "FFFFFF"

        for col in range(1, 6):
            cell = ws.cell(r, col)
            cell.border = border_thin()
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if col in [1, 5] else "left",
                                       wrap_text=True)
            if col == 1:
                cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
                cell.fill = fill(color)
            elif col == 2:
                cell.font = Font(name="Arial", bold=True, size=9, color="0D1B2A")
                cell.fill = fill(bg)
            elif col == 3:
                cell.font = Font(name="Arial", size=9, color="374151")
                cell.fill = fill(bg)
            elif col == 4:
                cell.font = Font(name="Arial", size=8, color="64748B", italic=(idx > 0))
                cell.fill = fill(bg)
            else:
                cell.font = Font(name="Arial", bold=True, size=9,
                                 color="1D4ED8" if lang == "EN" else "C0392B" if lang == "KO" else "5A3E7A")
                cell.fill = fill(bg)

# 열 너비
col_widths = [26, 24, 20, 42, 8]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── 세 번째 시트: 카테고리 설명
ws2 = wb.create_sheet("카테고리·사용방법")
ws2.merge_cells("A1:D1")
ws2["A1"] = "카테고리 분류 및 키워드 사용 방법"
ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
ws2["A1"].fill = fill("0D1B2A")
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

categories = [
    ("military",        "군사",       "C0392B", "전투·공습·군사 작전 관련 기사"),
    ("diplomacy",       "외교",       "1D4ED8", "협상·제재·국제관계 관련 기사"),
    ("energy",          "에너지",     "D97706", "유가·LNG·에너지 공급 관련 기사"),
    ("economy",         "경제",       "1A6B4A", "물가·공급망·성장률 관련 기사"),
    ("humanitarian",    "인도주의",   "5A3E7A", "민간 피해·난민·인권 관련 기사"),
    ("nuclear",         "핵",         "6B7280", "이란 핵 프로그램 관련 기사"),
    ("korea",           "한국",       "0D1B2A", "한국 경제·정책 직접 영향 기사"),
    ("paradigm",        "패러다임",   "2E4E7E", "에너지·무역 구조 재편 신호 기사"),
    ("country_response","각국 대응",  "1E6B8A", "각국 정책·외교 대응 현황 기사"),
]

ws2.append([])
ws2.append(["카테고리 코드", "카테고리명", "설명", "수원시 활용"])
hr2 = ws2.max_row
for col in range(1, 5):
    c = ws2.cell(hr2, col)
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = fill("1B3A5C")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin()
ws2.row_dimensions[hr2].height = 20

usages = [
    "전황 판단, 호르무즈 봉쇄 시나리오 평가",
    "제재 수위, 협상 가능성 평가",
    "도시가스·전기요금 인상 압력 예측",
    "물가·소상공인·고용 영향 분석",
    "취약계층 지원 정책 근거",
    "이란 핵 협상 진행 상황 모니터링",
    "수원시 직접 영향 — 삼성전자 공급망·LNG 수입",
    "중장기 에너지·무역 구조 재편 시나리오",
    "각국 대응 매트릭스 업데이트",
]

for i, ((code, name, color, desc), usage) in enumerate(zip(categories, usages)):
    ws2.append([code, name, desc, usage])
    r = ws2.max_row
    ws2.row_dimensions[r].height = 20
    bg = "F8FAFC" if i % 2 == 0 else "FFFFFF"
    for col in range(1, 5):
        cell = ws2.cell(r, col)
        cell.border = border_thin()
        cell.alignment = Alignment(vertical="center", wrap_text=True,
                                   horizontal="center" if col <= 2 else "left")
        if col == 1:
            cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            cell.fill = fill(color)
        elif col == 2:
            cell.font = Font(name="Arial", bold=True, size=9, color="0D1B2A")
            cell.fill = fill(bg)
        else:
            cell.font = Font(name="Arial", size=9, color="374151")
            cell.fill = fill(bg)

for i, w in enumerate([18, 14, 32, 36], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ── 저장
out = r"C:\Users\user\Desktop\iran_news_agent_final\iran_final\이란전쟁_에이전트_소스목록.xlsx"
wb.save(out)
print(f"완료: {out}")
